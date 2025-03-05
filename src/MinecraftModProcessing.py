import requests
import tldextract
import re
import webbrowser
import os
from bs4 import BeautifulSoup
import chardet

import pandas as pd
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

import datetime
import time
import tqdm
import openpyxl
from openpyxl import load_workbook
import lxml
import numpy as np

from MinecraftModScraping import *

def process(b1: Bundle) -> None:
    """
    """
    #if we do not scrape more data
    if not b1.mode:
        while True:
            ans = str(input("What mode of data collected would you like to process (general/detail): ").strip().lower())

            if ans in ['general', 'detail']:
                b1.mode = ans
                print(f"Mode selected: {b1.mode}")
                break

            else:
                print("Invalid response. Please enter (general) or (detail).")
                
                
    #combine all webpages into one one MASTER page for all subdirectories of a mode
    create_aggregated_file(b1)
    
    #Scan raw scraped data for encoding type
    encoding = detect_encoding(b1)
    b1.encoding = encoding
    
    #access the one MASTER webpage and put the desired data into a pandas DataFrame
    combinedDF = convert_html_to_df(b1)
    
    
    
    
    

    if b1.mode == 'general':
        
        #adjust the values(dates ,general cleaning, etc)
        combinedDF = clean_df_general(combinedDF)
    
    else:
        
        combinedDF = clean_df_detail(combinedDF)
        
        
    write_to_excel(combinedDF, b1)
    

def write_to_excel(combinedDF: pd.DataFrame, b1:Bundle, sheetName:str):
    """
    decription: 
            The purpose of this function is to write the results of the scraped data
            from curseforge.net to an excel file in an Output directory

    parameters: 
            combinedDF (pd.DataFrame): is the resulting dataframe of the script
            b1 (Bundle):

    returns:
        None
    """

    fileName = f'output.xlsx'

    directoryPath = os.path.join(b1.currentDirectory,
                                 b1.parentDataDirectory,
                                 b1.outputDataDirectory,
                                 b1.finalDataDirectory)
    
    if sheetName is None:
        sheetName = b1.mode

    try:
        if not os.path.exists(directoryPath):
            os.makedirs(directoryPath)

        filePath = os.path.join(directoryPath, fileName)
        if os.path.exists(filePath):
            try:
                
                # Load the existing workbook
                workbook = openpyxl.load_workbook(filePath)
                #print(f"Workbook loaded with sheetnames found: {workbook.sheetnames}")

                # if sheet exists already, let's replace it without deleting other sheets
                if b1.mode in workbook.sheetnames:
                    
                    with pd.ExcelWriter(filePath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        combinedDF.to_excel(writer, sheet_name=sheetName, index=False)
                
                # if it does not exist, create new sheet without deleting the others
                else:
                    with pd.ExcelWriter(filePath, engine='openpyxl', mode='a') as writer:
                        combinedDF.to_excel(writer, sheet_name=sheetName, index=False)
                        
            except Exception as e:
                raise f'Could not write to existing workbook: {str(e)}'
        else:
            
            #if the file did not exist at all create a new one
            with pd.ExcelWriter(filePath, engine='openpyxl', mode='w') as writer:
                combinedDF.to_excel(writer, sheet_name=sheetName, index=False)
    except Exception as e:
        raise "Could not output the results. Check Function 'write_to_excel()'"
    
    print(f"File: {os.path.basename(filePath)} written containing sheets {workbook.sheetnames}")

def clean_df_general(combinedDF: pd.DataFrame) -> pd.DataFrame:

    combinedDFColumns = combinedDF.columns

    #strip white space
    for i in range(len(combinedDFColumns)):
        combinedDF[combinedDFColumns[i]] = combinedDF[combinedDFColumns[i]].str.strip()

    #adjust mod_loader column
    combinedDF['mod_loader'] = combinedDF['mod_loader'].str.split().apply(lambda x: '|'.join([item.strip().replace(r'\+.', '') for item in x if len(item.strip()) > 1]))


    #adjust dates
    combinedDF['date_created'] = pd.to_datetime(combinedDF['date_created'], format='%b %d, %Y').dt.strftime('%d-%m-%Y')
    combinedDF['date_last_updated'] = pd.to_datetime(combinedDF['date_last_updated'], format='%b %d, %Y').dt.strftime('%d-%m-%Y')
    combinedDF['c_date_scraped'] = pd.to_datetime(combinedDF['c_date_scraped'], format='%d-%m-%Y').dt.strftime('%d-%m-%Y')

    # Sort the DataFrame by 'date_scraped' in ascending order
    combinedDF.sort_values('c_date_scraped', inplace=True)

    # Keep the newest entry for each title based on 'date_scraped'
    combinedDF['c_date_scraped'] = pd.to_datetime(combinedDF['c_date_scraped'], format='%d-%m-%Y')  # Ensure correct datetime format
    combinedDF = combinedDF.loc[combinedDF.groupby('title')['c_date_scraped'].idxmax()]
    combinedDF['c_date_scraped'] = combinedDF['c_date_scraped'].dt.strftime('%d-%m-%Y')

    # Reset index after grouping
    combinedDF.reset_index(drop=True, inplace=True)

    


    return combinedDF

def convert_html_to_df(b1: Bundle):
    """
    """
    #get parentDirectory path
    directory = os.path.join(b1.currentDirectory,
                             b1.parentDataDirectory,
                             b1.outputDataDirectory,
                             b1.rawDataDirectory,
                             b1.mode)

    #get all items from parentDirectory and filter it down to only subdirectories
    items = os.listdir(directory)
    folders = [os.path.join(directory, item) for item in items if os.path.isdir(os.path.join(directory, item))]
    
    #Intantiate some vars
    master = 'MASTER'
    finalDF = pd.DataFrame()
    dfs = []

    #get every curseforge.com scraped data's master file
    for folder in folders:
        files = sorted(os.listdir(folder), reverse=False)
        files = [os.path.join(folder, file) for file in files]
        masterFile = [s for s in files if master in s]

        if not masterFile:
            print(f"{folder}: did not have a MASTER file and thus will not be processed")
            continue
        
        print(f"Reading in file: {os.path.basename(masterFile[0])} using {b1.encoding[os.path.basename(masterFile[0])]} encoding") #perform slice because masterFile is a list, we want the file
        with open(masterFile[0], 'r', encoding='utf-8') as file: #encoding=b1.encoding[os.path.basename(masterFile[0])]
            content = file.read()
        soup = BeautifulSoup(content, 'lxml')
        print(f"File: {os.path.basename(masterFile[0])} loaded.")
        #print(soup)
        
            
        if b1.mode == 'general':

            
            cards = soup.find_all('div', class_='project-card')

            dataList = []
            progress_bar = tqdm.tqdm(cards, desc=f"Building DataFrame: {os.path.basename(folder)}", colour='#7B64C2')

            for card in progress_bar:
                a = card.find('a', class_='name')
                b = card.find('a', class_='author-name')
                c = card.find('p', class_='description')
                d = card.find('li', class_='detail-flavor')
                e = card.find('li', class_='detail-game-version')
                f = card.find('li', class_='detail-size')
                g = card.find('li', class_='detail-created')
                h = card.find('li', class_='detail-updated')
                j = card.find('li', class_='detail-downloads')
                k = card.find('ul', class_='categories')
                kCategories = [li.text.strip() for li in k.find_all('li')] if k else []

                dataList.append({
                    'title': '' if a is None else a.get('title')[a.get('title').find('Go to') + 6:a.get('title').find('Project Page')],
                    'author': '' if b is None else b.get('href')[b.get('href').rfind('/') + 1:],
                    'description': '' if c is None else c.text,
                    'mod_loader': '' if d is None else d.text,
                    'version': '' if e is None else e.text,
                    'file_size': '' if f is None else f.text,
                    'date_created': '' if g is None else g.text,
                    'date_last_updated': '' if h is None else h.text,
                    'download_count': '' if j is None else j.text,
                    'mod_category': '' if not kCategories else ' | '.join(kCategories),
                    'c_link_extension': '' if a is None else a.get('href'),
                    'c_date_scraped': folder[folder.rfind('--') + 2:]
                })

            df = pd.DataFrame(dataList)
            df.drop_duplicates(inplace=True, keep='first')
            dfs.append(df)
            
        elif b1.mode == 'detail':
            cards = soup.find_all('div', class_='files-table')
            print(f"Tables found: {len(cards)}")
                        
            progressBar = tqdm.tqdm(cards, desc=f"Building DataFrame: {os.path.basename(folder)}", colour='#7B64C2')
            
            for card in progressBar:
                #headerRow = card.find('div', class_='header-row')
                #columns = [item['title'] for item in headerRow.find_all('div')]
                data=[]
                rows=card.find_all('div', class_='file-row')
                #print(f"card: {card.find('a', class_="file-row-details")['href']} has len of rows: {len(rows)}")
                #tqdm.tqdm.write(f"\rCard: {card.find('a', class_='file-row-details')['href']} has len of rows: {len(rows)}")

                
                for index, row in enumerate(rows):
                    L = [elem.text.strip() for elem in row.find_all('span', class_=False)] + [elem.text.strip() for elem in row.find_all('div', class_='tooltip tooltip-small')]
                    L = list(set(L))
            

                    rowData = {
                            b1.filesColumns[b1.filesColumns.index('Type')]: row.find('span', class_='channel-tag release').text.strip() if row.find('span', class_='channel-tag release') 
                                                                        else row.find('span', class_='channel-tag alpha').text.strip() if row.find('span', class_='channel-tag alpha') 
                                                                        else row.find('span', class_='channel-tag beta').text.strip() if row.find('span', class_='channel-tag beta') 
                                                                        else 'No data recorded on website'
                                                                        ,
                            b1.filesColumns[b1.filesColumns.index('Name')]: row.find('span', class_='name')['title'].strip() if row.find('span', class_='name') else 'No data recorded on website',
                            b1.filesColumns[b1.filesColumns.index('Uploaded')]: next((elem for elem in L if b1.is_valid_date_format(b1.dateFormats, elem)), 'No data recorded on website'),
                            b1.filesColumns[b1.filesColumns.index('Size')]: next((elem for elem in L if re.search(r'\d+(\.\d+)?\s?(KB|MB|GB)', elem)), 'No data recorded on website'),
                            b1.filesColumns[b1.filesColumns.index('Game Ver.')]: ', '.join(sorted(set(re.findall(r'(?<!\S)\d+\.\d+(?:\.\d+)*(?!\S)', ' '.join([elem for elem in L if not re.search(r'[a-zA-Z]', elem)]))))) or 'No data recorded on website',
                            b1.filesColumns[b1.filesColumns.index('Mod Loaders')]: ', '.join(sorted(set(elem.strip() for elem in L if not re.search(r'\d', elem) and not re.fullmatch(r'release|alpha|beta', elem, re.IGNORECASE)))) or 'No data recorded on website',
                            b1.filesColumns[b1.filesColumns.index('Downloads')]: next((int(elem.replace(',', '')) for elem in L if elem.replace(',', '').isdigit()), 'No data recorded on website'),
                            b1.filesColumns[b1.filesColumns.index('c_link_extension')]: card.find('a', class_='file-row-details')['href'].split('/files')[0].strip().lower(),
                            b1.filesColumns[b1.filesColumns.index('c_date_scraped_detail')]: os.path.splitext(masterFile[0].split('MASTER')[-1])[0],
                            b1.filesColumns[b1.filesColumns.index('c_download_link')] : str(row.find('a', class_='file-row-details')['href'].strip().lower())
                        }
        
                    
                    data.append(rowData)
                    
                df = pd.DataFrame(data, columns=b1.filesColumns)
                df = df.drop_duplicates(keep='first')
                dfs.append(df)
                    
                    
                
            
    if dfs and len(dfs) > 0:      
        finalDF = pd.concat(dfs, ignore_index=True)

        return finalDF

def create_aggregated_file(b1: Bundle):
    """
    description: 
            Per subdirectory, create a master file, which is a file that contains the contents of 
            all files within that subdirectory. This function accounts for subdirectories that already
            have a master file within it, and skips it to avoid compiling content that already exists.

    parameters:
            None

    returns:

            None
    """
    #get current directory with cachedData
    directory = os.path.join(os.getcwd(), 
                             b1.parentDataDirectory,
                             b1.outputDataDirectory,
                             b1.rawDataDirectory,
                             b1.mode)

    #get everything in the parent data directory
    items = os.listdir(directory)

    #get all subdirectories inside the parent data directory
    subdirectories = [os.path.join(directory, item) for item in items if os.path.isdir(os.path.join(directory, item))]


    #Check if the aggregated file already exists (if it exists, will skip recompilation)
        #if it contains the word MASTER-subdirectory and is the largest file add to set
        # we will remove the set from the subdirectories list
    aggregatedCheck = set()
    for index, subdirectory in enumerate(subdirectories):

        contents = set(os.listdir(subdirectory))
        #L = set(file for file in contents if f'MASTER{subdirectory.split('.com--')[1]}' in file)
        L = [
            max(
                (file for file in contents if f'MASTER{subdirectory.split(".com--")[1]}' in file),
                key=lambda f: os.path.getsize(os.path.join(subdirectory, f)),
                default=None
            )
        ]

        # Ensure L is an empty list if no file was found
        L = [] if L == [None] else L


        if len(L) > 0:
            aggregatedCheck.add(subdirectory)
        
        #remove the directories with Master already in it
    for index, subdirectory in enumerate(aggregatedCheck):
        
        if subdirectory in subdirectories:

            subdirectories.remove(subdirectory)
            print(f"{subdirectory.split(b1.rawDataDirectory)[-1]} already contained a master file")
            
    # if there are subdirectories that have not been compiled into a master file create a master
    # file per subdirectory
    if len(subdirectories) > 0:

        iterationCount = 0
        for index, subdirectory in enumerate(subdirectories):
            folder = os.listdir(subdirectory)
            iterationCount = iterationCount + len(subdirectory)

        progressBar = tqdm.tqdm(subdirectories, total=iterationCount, colour='#7B64C2')

        for subdirectory in subdirectories:

            #get every file in the folder
            L = sorted(os.listdir(subdirectory), reverse=False)
            L = [os.path.join(directory, subdirectory, file) for file in L if file.endswith('.html')]

            

            # Update the description with the current item
            progressBar.set_description(f"Buidling Master File: {subdirectory[subdirectory.rfind('/') + 1:]}")
            
            #get the file path of the last file in the folder
            master = ''
            master = L[-1]
            master = master.replace('DATA', 'MASTER')
            print(f"In subdirectory: {subdirectory} | Master file: {master}")
            
            #loop through every file in a folder
            
            with open(master, 'w', encoding='utf-8') as masterFile:
                masterFile.write(BeautifulSoup("<html>\n<body>", 'html.parser').prettify())
                
                for i in range(len(L)):

                    #read the file and compile all the files in a folder into one big string
                    try:
                        if L[i] == master:
                            pass
                        else:
                            with open(L[i], 'r', encoding='utf-8', errors='replace') as file:
                                content = file.read()
                                soup = BeautifulSoup(content, 'html.parser')
                                masterFile.write(f"<!-- Start of {os.path.basename(L[i])} -->\n")
                                masterFile.write(soup.prettify().encode('utf-8', 'ignore').decode('utf-8'))
                                #masterFile.write(soup.prettify())
                                masterFile.write(f"<!-- End of {os.path.basename(L[i])} -->\n")
                                if i%100 == 0:
                                    progressBar.update(1)
                    except Exception as e:
                        raise e
                    
                masterFile.write(BeautifulSoup("</body>\n</html>", 'html.parser').prettify())
                
        
        progressBar.close()
    else:

        print('All existing subdirectories contain a master file. Skipping compilation process.')
        
def clean_df_detail(combinedDF: pd.DataFrame) -> pd.DataFrame:
    """
    """
    
    #Adjust column names. lowercase and snakecase
    combinedDF.columns = [re.sub(r'[^a-z0-9_\s]', '', col.lower().strip()).replace(' ', '_') for col in combinedDF.columns]
    
    #Strip Columns
    for x in combinedDF.columns:
        
        #strip strings
        if combinedDF[x].dtype == 'object':
            combinedDF[x] = combinedDF[x].str.strip()
    
    #adjust logic   
        #Make them into whole words   
    combinedDF['type'] = np.where(combinedDF['type'] == 'B', "Beta",
                              np.where(combinedDF['type'] == 'R', "Release",
                                       np.where(combinedDF['type'] == 'A', "Alpha",
                                                "No data recorded on website")))

        #keep the date format consistent
    combinedDF['uploaded'] = pd.to_datetime(combinedDF['uploaded'], format="%b %d, %Y").dt.strftime("%m-%d-%Y")
        #create a column so we can select what mods we want
    combinedDF['c_selected'] = ''
        #clean up mod loaders, the oringinal calculation does not work all the way
    combinedDF['mod_loaders'] = combinedDF['mod_loaders'].apply(
    lambda x: ', '.join(sorted(set(i.strip() for i in re.split(r',| ', x) if i.strip())))
    )
        #infer a mod loader if they left it out when uploading the file
        ## Filter only rows where mod_loaders is not "No data recorded on website"
    validLoaders = combinedDF[combinedDF['mod_loaders'] != 'No, data, on, recorded, website']
    singleLoaderMap = validLoaders.groupby('c_link_extension')['mod_loaders'].nunique() == 1

    # Get the actual mod loader values for those groups
    singleLoaderValues = validLoaders.groupby('c_link_extension')['mod_loaders'].first()

    # Update the original DataFrame
    combinedDF['mod_loaders'] = combinedDF.apply(
        lambda row: singleLoaderValues[row['c_link_extension']]
        if row['c_link_extension'] in singleLoaderValues.index 
        and row['mod_loaders'] == 'No, data, on, recorded, website'
        else row['mod_loaders'],
        axis=1
    )


    
    
    
    return combinedDF            
        

def detect_encoding(b1: Bundle) -> str:
    """
    description: the purpose of this function is to detect the encoding of a master file. if the master file is not present
                we will get an encoding of every file in the directory we are trying to read
                
    params:
        b1: Bundle
        
    returns:
        encodings: dict
            a dictionary containing all the encodings
    """
    
    basePath = os.path.join(
            b1.currentDirectory,
            b1.parentDataDirectory,
            b1.outputDataDirectory,
            b1.rawDataDirectory,
            b1.mode 
        )
    encodings = {}
    #get all directories in the folders we want to process
    directories = [fold for fold in os.listdir(basePath) if os.path.isdir(os.path.join(basePath, fold))]
    
    
    for idx, directory in enumerate(directories):
        
        checkMaster = [file for file in os.listdir(os.path.join(basePath, directory)) if file.endswith('.html') and 'MASTER' in file]
        
        #if there is a master file, only focus on the encoding of that master file
        if checkMaster:
            
            with open(os.path.join(basePath, directory, checkMaster[0]), 'rb') as masterFile:
                rawData = masterFile.read(100**2)
            
            encodingInfo = chardet.detect(rawData)
            detectedEncoding = encodingInfo.get('encoding', 'uknown')
            encodings[checkMaster[0]] = detectedEncoding
        
        #if there is not a master file lets get the encoding of all files with in the directory
        else:
            
            files = [file for file in os.listdir(directory) if file.endswith('.html')]
            encodes = []
            
            for idxy, file in enumerate(files):
                
                with open(os.path.join(basePath, directory, file), 'rb') as shardFile:
                    rawData = shardFile.read()
                    
                encodingInfo = chardet.detect(rawData)
                detectedEncoding = encodingInfo.get('encoding', 'uknonwn')
                encodes.append(detectedEncoding)
            
               
            encodings[directory] = list(set(encodes))
            
    
    return encodings
                    
